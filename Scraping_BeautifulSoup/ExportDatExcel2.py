# -*- coding: utf-8 -*-
"""
Created on Mon Apr  5 22:34:04 2021

@author: MartinXD
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


def tamanioColumna(ruta ,nombre):
    
    file_excel = pd.read_excel(ruta)
    
    values = file_excel[nombre].values
    
    return len(values)

def leerColumna(ruta ,nombre):
    
    file_excel = pd.read_excel(ruta)
    
    values = file_excel[nombre].values
    
    return values

def leerExcel(ruta):
    
    file_excel = pd.read_excel(ruta)

    columnas = ['Productos', 'Jumbo', 'Dia', 'Coto', 'Walmart', 'Fecha']
    df_seleccionados = file_excel[columnas]
    
    print(df_seleccionados)

def crearExcel(ruta):
    
    wb = Workbook()
    
    filesheet = ruta
    
    wb.save(filesheet)
    
def cargarIniciales(ruta, datos):
    
    filesheet = ruta
    
    wb = load_workbook(filesheet)
    
    sheet = wb.active
    
    columna = 1
    fila = 1
    
    for dato in datos:
        sheet.cell(fila, columna, dato)
        columna+=1
        
    wb.save(filesheet)

def cargarAExcel(ruta ,c ,datos):
    
    filesheet = ruta
    
    wb = load_workbook(filesheet)
    
    sheet = wb.active
    
    now = datetime.now()

    fecha = now.date()
    
    columnaClave = 1
    columna = c  #variable que maneja las columnas
    
    for k,v in datos.items():
        fila = 1  #variable que maneja las filas
        keyColumn = leerColumna(ruta ,'Productos')
        if k in keyColumn:
            for c in keyColumn:
                fila+=1
                if c == k:
                    sheet.cell(fila, columna, v)
                    wb.save(filesheet)
        else:
            filaVacia = tamanioColumna(ruta ,'Productos') + 2
            sheet.cell(filaVacia, columnaClave, k)
            sheet.cell(filaVacia, columna, v)
            wb.save(filesheet)
            
    sheet.cell(2, 6, fecha)
    
    wb.save(filesheet)
#--------------------Main--------------------------
"""
iniciales = ["Colores", "Ingles", "Frances", "Portugues", "Italiano", "Fecha"]
coloresIngles = {"Rojo":"Red", "Verde":"Green", "Azul":"Blue"}
coloresFrances = {"Rojo":"Rouge", "Verde":"Vert", "Azul":"Bleu"}
coloresPortugues = {"Rojo":"vermelho", "Amarillo":"amarelo" ,"Verde":"verde", "Azul":"azul"}
coloresItaliano = {"Rojo":"rosso", "Amarillo":"giallo" ,"Verde":"verde", "Naranja":"arancione" ,"Azul":"blu"}

#now = datetime.now()

#fecha = now.date()

#crearExcel()

cargarAExcel(5, coloresItaliano)

#leerColumna('Ingles')

#tamanioColumna('Colores')

leerExcel()
"""

